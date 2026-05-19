"""
Gera planilha XLSX comercial com 6 tipos de aba para o time de vendas B2B.

Função pública: build_workbook(cursos) -> bytes
"""

import io
import re
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Estilos compartilhados ──────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
_LINK_FONT = Font(name="Calibri", size=11, color="2563EB", underline="single")
_BANNER_FONT = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
_BANNER_ROW_HEIGHT = 36

_THIN = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
_ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

_PALETA = [
    "0EA5E9", "F59E0B", "10B981", "8B5CF6", "EF4444",
    "EC4899", "06B6D4", "84CC16", "F97316", "6366F1",
    "14B8A6", "F43F5E", "A855F7", "22C55E", "EAB308",
]


def _cor_estavel(nome: str) -> str:
    """Cor consistente entre execuções (hash determinístico do nome)."""
    h = 0
    for ch in nome:
        h = (h * 31 + ord(ch)) & 0xFFFFFFFF
    return _PALETA[h % len(_PALETA)]


def _sanitizar_nome_aba(nome: str) -> str:
    """Excel limita a 31 chars e proíbe : \\ / ? * [ ]"""
    nome = re.sub(r"[\\/*?:\[\]]", "", nome)
    return nome[:31]


def _ajustar_larguras(ws, larguras: dict):
    for col_letter, largura in larguras.items():
        ws.column_dimensions[col_letter].width = largura


def _aplicar_banner(ws, row: int, max_col: int, texto: str, cor_hex: str):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.value = texto
    cell.fill = PatternFill("solid", fgColor=cor_hex)
    cell.font = _BANNER_FONT
    cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[row].height = _BANNER_ROW_HEIGHT


def _aplicar_header(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER


def _nomes_instrutores(curso: dict) -> str:
    return ", ".join(
        i.get("nome", "") for i in curso.get("instrutores", []) or [] if i.get("nome")
    )


def _nomes_carreiras(curso: dict) -> str:
    return ", ".join(
        c.get("titulo", "") for c in curso.get("carreiras", []) or [] if c.get("titulo")
    )


def _qtd_habilidades(curso: dict) -> int:
    return sum(len(c.get("habilidades", []) or []) for c in curso.get("competencias", []) or [])


def _nome_or_str(v) -> str:
    """Aceita str ou dict {'nome': ...} e devolve string. Outros tipos → ''."""
    if isinstance(v, dict):
        return v.get("nome", "") or ""
    if isinstance(v, str):
        return v
    return ""


def _categorias_do_curso(curso: dict) -> set[str]:
    """Coleta categoria principal + lista categorias[] (cada item pode ser str ou dict)."""
    cats: set[str] = set()
    nome_principal = _nome_or_str(curso.get("categoria"))
    if nome_principal:
        cats.add(nome_principal)
    for c in curso.get("categorias", []) or []:
        nome = _nome_or_str(c)
        if nome:
            cats.add(nome)
    return cats


# ─── Builders por aba ────────────────────────────────────────────────────────


def _build_capa_indice(wb: Workbook, cursos: list[dict]) -> None:
    ws = wb.create_sheet("📑 Capa & Índice")
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 20

    ws["A1"] = "Catálogo Alura — Versão Comercial"
    ws["A1"].font = Font(name="Calibri", size=24, bold=True, color="1E3A8A")
    ws.row_dimensions[1].height = 40

    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="64748B")

    categorias = {c for curso in cursos for c in _categorias_do_curso(curso)}
    carreiras = {
        car.get("titulo")
        for curso in cursos
        for car in (curso.get("carreiras", []) or [])
        if car.get("titulo")
    }

    stats = [
        ("Total de cursos:", len(cursos)),
        ("Categorias:", len(categorias)),
        ("Carreiras:", len(carreiras)),
    ]
    for i, (label, valor) in enumerate(stats, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=valor)

    ws["A8"] = "Abas neste arquivo"
    ws["A8"].font = Font(bold=True, size=14, color="1E3A8A")

    abas_doc = [
        ("🎯 Catálogo Geral", "1 linha por curso. Use Ctrl+F e os filtros do header para buscar."),
        ("📂 Por Categoria (uma aba por área)", "Cursos da categoria com competências e habilidades detalhadas em células mescladas."),
        ("🛤️ Por Carreira", "Tabela flat — 1 linha por (carreira × curso). Filtre por Carreira + Step + Categoria para responder 'que cursos compõem o nível Intermediário de Front-End?'."),
        ("🧠 Por Competência", "Tabela flat — 1 linha por (competência × habilidade × curso). Filtre combinando Competência + Categoria + Carreira para queries como 'cursos com competência X em Data Science'."),
    ]
    row = 9
    for nome, desc in abas_doc:
        ws.cell(row=row, column=1, value=nome).font = Font(bold=True, color="0F172A")
        ws.cell(row=row + 1, column=1, value=desc).alignment = _ALIGN_LEFT_TOP
        ws.cell(row=row + 1, column=1).font = Font(size=10, color="475569")
        row += 3


def _build_catalogo_geral(wb: Workbook, cursos: list[dict]) -> None:
    ws = wb.create_sheet("🎯 Catálogo Geral")
    headers = [
        "ID", "Nome", "Link", "Categoria", "Subcategoria", "Carreiras",
        "Instrutores", "Carga horária (h)", "Qtd aulas", "Min de vídeo",
        "Qtd alunos", "Nota", "Última atualização", "Público-alvo",
        "Resumo", "Nota do comercial",
    ]
    _aplicar_header(ws, 1, headers)

    for i, curso in enumerate(cursos, start=2):
        link = curso.get("link", "") or ""
        data_atu = curso.get("data_atualizacao") or ""
        if isinstance(data_atu, str):
            data_atu = data_atu[:10]
        row_data = [
            curso.get("course_id"),
            curso.get("nome", ""),
            link,
            _nome_or_str(curso.get("categoria")),
            _nome_or_str(curso.get("subcategoria")),
            _nomes_carreiras(curso),
            _nomes_instrutores(curso),
            curso.get("carga_horaria"),
            curso.get("quantidade_aulas"),
            curso.get("minutos_video"),
            curso.get("quantidade_alunos"),
            curso.get("nota"),
            data_atu,
            curso.get("publico_alvo", "") or "",
            (curso.get("metadescription") or "")[:300],
            "",
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = _ALIGN_LEFT_TOP
            cell.border = _BORDER
            if i % 2 == 0:
                cell.fill = _ZEBRA_FILL

        if link:
            link_cell = ws.cell(row=i, column=3)
            link_cell.hyperlink = link
            link_cell.font = _LINK_FONT

    ws.freeze_panes = "B2"
    last_row = len(cursos) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    _ajustar_larguras(ws, {
        "A": 8, "B": 50, "C": 40, "D": 22, "E": 22, "F": 30,
        "G": 30, "H": 14, "I": 10, "J": 12, "K": 12, "L": 8,
        "M": 14, "N": 30, "O": 60, "P": 25,
    })


def _build_abas_por_categoria(wb: Workbook, cursos: list[dict]) -> None:
    por_cat: dict[str, list[dict]] = defaultdict(list)
    for curso in cursos:
        for cat in _categorias_do_curso(curso):
            por_cat[cat].append(curso)

    for categoria in sorted(por_cat.keys()):
        cursos_cat = por_cat[categoria]
        ws = wb.create_sheet(_sanitizar_nome_aba(f"📂 {categoria}"))
        cor = _cor_estavel(categoria)

        _aplicar_banner(ws, 1, 6, f"{categoria} — {len(cursos_cat)} cursos", cor)
        _aplicar_header(ws, 3, ["ID", "Nome do curso", "Link", "Carga (h)", "Competência", "Habilidade"])

        row = 4
        for curso in cursos_cat:
            comps = curso.get("competencias", []) or []
            start_row = row
            link = curso.get("link", "") or ""

            # Preenche colunas 1-4 na primeira linha do bloco do curso
            valores_curso = [
                curso.get("course_id"),
                curso.get("nome", ""),
                link,
                curso.get("carga_horaria"),
            ]
            for col, v in enumerate(valores_curso, start=1):
                cell = ws.cell(row=start_row, column=col, value=v)
                cell.alignment = _ALIGN_LEFT_CENTER
                cell.border = _BORDER
            if link:
                link_cell = ws.cell(row=start_row, column=3)
                link_cell.hyperlink = link
                link_cell.font = _LINK_FONT

            if not comps:
                ws.cell(row=row, column=5, value="(não classificado)").alignment = _ALIGN_LEFT_TOP
                ws.cell(row=row, column=5).border = _BORDER
                ws.cell(row=row, column=6, value="").border = _BORDER
                row += 1
            else:
                for comp in comps:
                    habs = comp.get("habilidades", []) or []
                    n_habs = max(1, len(habs))
                    comp_label = f"{comp.get('codigo_competencia', '')} — {comp.get('nome_competencia', '')}"

                    if n_habs > 1:
                        ws.merge_cells(start_row=row, start_column=5, end_row=row + n_habs - 1, end_column=5)
                    cell = ws.cell(row=row, column=5, value=comp_label)
                    cell.alignment = _ALIGN_LEFT_CENTER
                    cell.border = _BORDER

                    if habs:
                        for hab in habs:
                            hab_label = f"{hab.get('codigo_habilidade', '')} — {hab.get('nome_habilidade', '')}"
                            c = ws.cell(row=row, column=6, value=hab_label)
                            c.alignment = _ALIGN_LEFT_TOP
                            c.border = _BORDER
                            row += 1
                    else:
                        ws.cell(row=row, column=6, value="").border = _BORDER
                        row += 1

            # Mesclar curso (cols 1-4) se ocupou mais de 1 linha
            end_row = row - 1
            if end_row > start_row:
                for col in (1, 2, 3, 4):
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

        ws.freeze_panes = "A4"
        _ajustar_larguras(ws, {"A": 8, "B": 45, "C": 35, "D": 10, "E": 50, "F": 50})


def _build_aba_por_carreira(wb: Workbook, cursos: list[dict]) -> None:
    """
    Tabela flat — uma linha por (carreira × curso).
    Permite filtrar por Carreira + Step + Categoria para responder perguntas como:
    'que cursos compõem o nível Intermediário de Front-End React?'
    """
    ws = wb.create_sheet("🛤️ Por Carreira")
    headers = [
        "Carreira", "Step", "Posição",
        "ID", "Curso", "Link",
        "Categoria", "Subcategoria",
        "Carga (h)", "Qtd aulas", "Min de vídeo", "Nota",
    ]
    _aplicar_header(ws, 1, headers)

    linhas: list[tuple] = []
    for curso in cursos:
        for carreira in curso.get("carreiras", []) or []:
            titulo = carreira.get("titulo")
            if not titulo:
                continue
            linhas.append((
                titulo,
                carreira.get("step_titulo", ""),
                carreira.get("step_position", 0),
                curso.get("course_id"),
                curso.get("nome", ""),
                curso.get("link", "") or "",
                _nome_or_str(curso.get("categoria")),
                _nome_or_str(curso.get("subcategoria")),
                curso.get("carga_horaria"),
                curso.get("quantidade_aulas"),
                curso.get("minutos_video"),
                curso.get("nota"),
            ))

    if not linhas:
        ws.cell(row=2, column=1, value="Nenhuma carreira cadastrada no banco.")
        return

    # Ordena: Carreira → Posição → Nome do curso
    linhas.sort(key=lambda r: (r[0], r[2], r[4]))

    for i, valores in enumerate(linhas, start=2):
        for col, v in enumerate(valores, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = _ALIGN_LEFT_TOP
            cell.border = _BORDER
            if i % 2 == 0:
                cell.fill = _ZEBRA_FILL
        link = valores[5]
        if link:
            link_cell = ws.cell(row=i, column=6)
            link_cell.hyperlink = link
            link_cell.font = _LINK_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(linhas) + 1}"
    _ajustar_larguras(ws, {
        "A": 30, "B": 22, "C": 10,
        "D": 8, "E": 50, "F": 35,
        "G": 22, "H": 22,
        "I": 10, "J": 10, "K": 12, "L": 8,
    })


def _build_aba_por_competencia(wb: Workbook, cursos: list[dict]) -> None:
    """
    Tabela flat — uma linha por (competência × habilidade × curso).
    Permite filtros combinados como:
      - 'cursos com competência X na carreira Front-End React'
      - 'cursos com competência Y na categoria Data Science'
      - 'todas as habilidades de uma competência específica'
    """
    ws = wb.create_sheet("🧠 Por Competência")
    headers = [
        "Cód. Competência", "Competência", "Descrição da Competência",
        "Cód. Habilidade", "Habilidade",
        "ID", "Curso", "Link",
        "Categoria", "Subcategoria", "Carreiras",
        "Carga (h)", "Nota",
    ]
    _aplicar_header(ws, 1, headers)

    linhas: list[tuple] = []
    for curso in cursos:
        for comp in curso.get("competencias", []) or []:
            cod_comp = comp.get("codigo_competencia", "")
            nome_comp = comp.get("nome_competencia", "")
            desc_comp = comp.get("descricao_competencia", "")
            habs = comp.get("habilidades", []) or []

            curso_meta = (
                curso.get("course_id"),
                curso.get("nome", ""),
                curso.get("link", "") or "",
                _nome_or_str(curso.get("categoria")),
                _nome_or_str(curso.get("subcategoria")),
                _nomes_carreiras(curso),
                curso.get("carga_horaria"),
                curso.get("nota"),
            )

            if not habs:
                linhas.append((cod_comp, nome_comp, desc_comp, "", "", *curso_meta))
            else:
                for hab in habs:
                    linhas.append((
                        cod_comp, nome_comp, desc_comp,
                        hab.get("codigo_habilidade", ""),
                        hab.get("nome_habilidade", ""),
                        *curso_meta,
                    ))

    if not linhas:
        ws.cell(row=2, column=1, value="Nenhuma competência encontrada nos cursos classificados.")
        return

    # Ordena: Competência → Habilidade → Carreira → Curso
    linhas.sort(key=lambda r: (r[0], r[3], r[10], r[6]))

    for i, valores in enumerate(linhas, start=2):
        for col, v in enumerate(valores, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.alignment = _ALIGN_LEFT_TOP
            cell.border = _BORDER
            if i % 2 == 0:
                cell.fill = _ZEBRA_FILL
        link = valores[7]
        if link:
            link_cell = ws.cell(row=i, column=8)
            link_cell.hyperlink = link
            link_cell.font = _LINK_FONT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(linhas) + 1}"
    _ajustar_larguras(ws, {
        "A": 14, "B": 35, "C": 50,
        "D": 14, "E": 35,
        "F": 8, "G": 45, "H": 35,
        "I": 22, "J": 22, "K": 30,
        "L": 10, "M": 8,
    })


# ─── Função pública ──────────────────────────────────────────────────────────


def build_workbook(cursos: list[dict]) -> bytes:
    """Monta o workbook completo e retorna os bytes do .xlsx."""
    wb = Workbook()
    wb.remove(wb.active)

    _build_capa_indice(wb, cursos)
    _build_catalogo_geral(wb, cursos)
    _build_abas_por_categoria(wb, cursos)
    _build_aba_por_carreira(wb, cursos)
    _build_aba_por_competencia(wb, cursos)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
